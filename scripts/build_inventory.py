from __future__ import annotations

import json
import re
import shutil
from pathlib import Path
from urllib.parse import quote

import qrcode
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "MasterInventory.xlsx"
SITE_SOURCE = ROOT / "site"
PHOTOS_SOURCE = ROOT / "photos"
BUILD = ROOT / "_site"

# This project is permanently pointed at this GitHub Pages site.
BASE_URL = "https://shammy33.github.io/PCS-INVENTORY"

EXPECTED = {
    "ITEM ID": "item_id",
    "ITEM NAME": "item_name",
    "CATEGORY": "category",
    "SERIAL NUMBER": "serial_number",
    "DESCRIPTION": "description",
    "BATTERY TYPE": "battery_type",
    "BATTERY WH": "battery_wh",
    "HAZARDOUS/RESTRICTED": "hazardous_restricted",
    "SHIPMENT LOCATION": "shipment_location",
    "NOTES": "notes",
    "PHOTOS": "photos_column",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def clean_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip().upper()


def clean_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_item_id(value: object) -> str:
    """Keep IDs as simple numbers: 1 -> 0001, 42 -> 0042, 1000 -> 1000."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = int(value)
        return str(number).zfill(4)
    text = str(value).strip()
    if text.isdigit():
        return text.zfill(4)
    return text


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(ws.max_row, 30) + 1):
        found = {}
        for cell in ws[row_number]:
            header = clean_header(cell.value)
            if header in EXPECTED:
                found[header] = cell.column
        if "ITEM ID" in found and "ITEM NAME" in found:
            return row_number, found
    raise ValueError("Could not find a header row containing ITEM ID and ITEM NAME.")


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(
            "MasterInventory.xlsx was not found in the repository root. "
            "Put your existing spreadsheet there and name it MasterInventory.xlsx."
        )

    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb.active
    header_row, columns = find_header_row(ws)

    missing = [name for name in EXPECTED if name not in columns]
    if missing:
        print("Warning: missing optional columns:", ", ".join(missing))

    items: dict[str, dict] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(row=row, column=columns["ITEM ID"]).value
        item_id = normalize_item_id(raw_id)
        if not item_id:
            continue
        if item_id in items:
            raise ValueError(f"Duplicate ITEM ID found: {item_id}")

        record = {}
        for sheet_header, key in EXPECTED.items():
            if key == "photos_column":
                continue
            col = columns.get(sheet_header)
            record[key] = clean_value(ws.cell(row=row, column=col).value) if col else ""
        record["item_id"] = item_id

        item_photo_dir = PHOTOS_SOURCE / item_id
        photos = []
        if item_photo_dir.is_dir():
            for photo in sorted(item_photo_dir.iterdir(), key=lambda p: p.name.lower()):
                if photo.is_file() and photo.suffix.lower() in IMAGE_EXTENSIONS:
                    photos.append(f"photos/{quote(item_id)}/{quote(photo.name)}")
        record["photos"] = photos
        items[item_id] = record

    if BUILD.exists():
        shutil.rmtree(BUILD)
    shutil.copytree(SITE_SOURCE, BUILD)
    if PHOTOS_SOURCE.exists():
        shutil.copytree(PHOTOS_SOURCE, BUILD / "photos", dirs_exist_ok=True)

    inventory_json = json.dumps(items, ensure_ascii=False, indent=2)
    (BUILD / "inventory.js").write_text(
        "window.INVENTORY = " + inventory_json + ";\n",
        encoding="utf-8",
    )
    (BUILD / ".nojekyll").write_text("", encoding="utf-8")

    # QR codes are included in the deployed site so each one can be opened/downloaded.
    qr_site_dir = BUILD / "qr-codes"
    qr_site_dir.mkdir(parents=True, exist_ok=True)
    for item_id in items:
        item_url = f"{BASE_URL}/?item={quote(item_id)}"
        img = qrcode.make(item_url)
        img.save(qr_site_dir / f"{item_id}.png")

    print(f"Built {len(items)} item records from {WORKBOOK.name}.")
    print(f"Site URL: {BASE_URL}/")
    print(f"Example item URL: {BASE_URL}/?item=0001")
    print(f"Example QR image: {BASE_URL}/qr-codes/0001.png")


if __name__ == "__main__":
    main()
