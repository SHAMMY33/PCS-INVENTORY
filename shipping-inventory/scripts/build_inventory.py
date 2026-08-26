from __future__ import annotations

import json
import os
import re
import shutil
from pathlib import Path
from urllib.parse import quote

import qrcode
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "MasterInventory.xlsx"
SITE_SOURCE = ROOT / "site"
PHOTOS_SOURCE = ROOT / "photos"
QR_OUTPUT = ROOT / "qr-codes"
BUILD = ROOT / "_site"

EXPECTED = {
    "ITEM ID": "item_id",
    "ITEM NAME": "item_name",
    "CATEGORY": "category",
    "SERIAL NUMBER": "serial_number",
    "DESCRIPTION": "description",
    "BATTERY TYPE": "battery_type",
    "BATTERY WH": "battery_wh",
    "HAZARDOUS/RESTRICTED": "hazardous_restricted",
    "SHIPMENT LOCATION": "shipment_location",
    "NOTES": "notes",
    "PHOTOS": "photos_column",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def clean_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip().upper()
    # Makes BATTERY Wh / BATTERY WH equivalent.
    return text


def clean_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(ws.max_row, 30) + 1):
        found = {}
        for cell in ws[row_number]:
            header = clean_header(cell.value)
            if header in EXPECTED:
                found[header] = cell.column
        if "ITEM ID" in found and "ITEM NAME" in found:
            return row_number, found
    raise ValueError("Could not find a header row containing ITEM ID and ITEM NAME.")


def get_base_url() -> str:
    override = os.getenv("INVENTORY_BASE_URL", "").strip().rstrip("/")
    if override:
        return override
    repo = os.getenv("GITHUB_REPOSITORY", "")
    if "/" in repo:
        owner, name = repo.split("/", 1)
        return f"https://{owner}.github.io/{name}"
    return "http://localhost:8000"


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(
            "MasterInventory.xlsx was not found in the repository root. "
            "Rename your master spreadsheet to MasterInventory.xlsx and upload it there."
        )

    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb.active
    header_row, columns = find_header_row(ws)

    missing = [name for name in EXPECTED if name not in columns]
    if missing:
        print("Warning: missing optional columns:", ", ".join(missing))

    items: dict[str, dict] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        raw_id = ws.cell(row=row, column=columns["ITEM ID"]).value
        if raw_id is None or str(raw_id).strip() == "":
            continue
        item_id = str(raw_id).strip().upper()
        if item_id in items:
            raise ValueError(f"Duplicate ITEM ID found: {item_id}")

        record = {}
        for sheet_header, key in EXPECTED.items():
            if key == "photos_column":
                continue
            col = columns.get(sheet_header)
            record[key] = clean_value(ws.cell(row=row, column=col).value) if col else ""
        record["item_id"] = item_id

        item_photo_dir = PHOTOS_SOURCE / item_id
        photos = []
        if item_photo_dir.is_dir():
            for photo in sorted(item_photo_dir.iterdir(), key=lambda p: p.name.lower()):
                if photo.is_file() and photo.suffix.lower() in IMAGE_EXTENSIONS:
                    photos.append(f"photos/{quote(item_id)}/{quote(photo.name)}")
        record["photos"] = photos
        items[item_id] = record

    if BUILD.exists():
        shutil.rmtree(BUILD)
    shutil.copytree(SITE_SOURCE, BUILD)
    if PHOTOS_SOURCE.exists():
        shutil.copytree(PHOTOS_SOURCE, BUILD / "photos", dirs_exist_ok=True)

    inventory_json = json.dumps(items, ensure_ascii=False, indent=2)
    (BUILD / "inventory.js").write_text(
        "window.INVENTORY = " + inventory_json + ";\n",
        encoding="utf-8",
    )
    (BUILD / ".nojekyll").write_text("", encoding="utf-8")

    QR_OUTPUT.mkdir(exist_ok=True)
    for old in QR_OUTPUT.glob("*.png"):
        old.unlink()

    base_url = get_base_url()
    for item_id in items:
        url = f"{base_url}/?item={quote(item_id)}"
        img = qrcode.make(url)
        img.save(QR_OUTPUT / f"{item_id}.png")

    print(f"Built {len(items)} item records from {WORKBOOK.name}.")
    print(f"QR base URL: {base_url}")


if __name__ == "__main__":
    main()
